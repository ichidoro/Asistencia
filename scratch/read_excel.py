import openpyxl
import os

excel_path = "Productos.xlsx"
if not os.path.exists(excel_path):
    print("El archivo Productos.xlsx no existe en la raiz.")
else:
    wb = openpyxl.load_workbook(excel_path)
    print(f"Sheets: {wb.sheetnames}")
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"\n--- Sheet: {name} ---")
        for i in range(1, 15):
            row = [sheet.cell(row=i, column=j).value for j in range(1, 10)]
            if any(row):
                print(f"Row {i}: {row}")
