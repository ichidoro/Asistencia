"""
BioAlba Scraper - VERSIÓN CON EXCEL
Descarga el Excel completo de usuarios (sin paginación)
"""

from typing import List, Dict, Any
from datetime import datetime
from loguru import logger
import traceback
import time as _t
import aiohttp
from bs4 import BeautifulSoup
import re
from io import BytesIO

# [FASE 2 OPT] python-calamine: parser Excel escrito en Rust puro.
# Benchmark vs openpyxl: 10-15× más rápido en lectura, sin parseo de DOM completo.
# Si no está instalado, cae al fallback openpyxl automáticamente.
try:
    from python_calamine import CalamineWorkbook
    _CALAMINE_AVAILABLE = True
except ImportError:
    from openpyxl import load_workbook as _openpyxl_load
    _CALAMINE_AVAILABLE = False
    logger.warning("⚠️ python-calamine no instalado — usando openpyxl (más lento). Instala: pip install python-calamine")

from backend.core.config import settings


class BioAlbaScraper:
    """Scraper para bioalba1.controlasistencia.cl usando exportación Excel.
    
    Optimización de sesión: La sesión HTTP y el login se reutilizan entre
    operaciones sucesivas. Solo se re-loguea si la sesión expiró o fue cerrada.
    Esto reduce de N logins a 1 por ciclo de vida de la instancia.
    """
    
    # Contador de context managers activos para saber cuándo cerrar la sesión
    _ref_count: int = 0
    
    def __init__(self):
        self.base_url = settings.CONTROL_ASISTENCIA_URL
        self.username = settings.CONTROL_ASISTENCIA_USER
        self.password = settings.CONTROL_ASISTENCIA_PASSWORD
        self.session = None
        self.is_logged_in = False
    
    async def ensure_session(self):
        """Garantiza que exista una sesión HTTP abierta. Lazy-init."""
        if self.session is None or self.session.closed:
            self.session = aiohttp.ClientSession()
            self.is_logged_in = False  # sesión nueva → requiere login
            logger.debug("🔌 Nueva sesión HTTP BioAlba creada")
    
    async def ensure_logged_in(self) -> bool:
        """Garantiza login activo, re-logueando solo si es necesario."""
        await self.ensure_session()
        if not self.is_logged_in:
            return await self.login()
        return True
    
    async def __aenter__(self):
        """Context manager entry — reutiliza sesión existente."""
        self._ref_count += 1
        await self.ensure_session()
        return self
    
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit — mantiene sesión abierta para reutilización.
        
        La sesión se cierra solo via close() explícito (ej: shutdown de app).
        Esto permite que operaciones secuenciales (preview → sync → marcaciones)
        reutilicen el mismo login sin re-autenticarse.
        """
        self._ref_count = max(0, self._ref_count - 1)
        # NO cerrar sesión aquí — se reutiliza en la siguiente operación
    
    async def close(self):
        """Cierre explícito para cleanup final."""
        if self.session and not self.session.closed:
            await self.session.close()
            self.is_logged_in = False
            self._ref_count = 0
    
    async def login(self) -> bool:
        """Login simplificado"""
        try:
            logger.info(f"🔐 Intentando login como {self.username}...")
            
            # Paso 1: GET a /login para obtener CSRF token
            async with self.session.get(f"{self.base_url}/login") as response:
                html = await response.text()
                soup = BeautifulSoup(html, 'html.parser')
                
                # Buscar CSRF token
                csrf_input = soup.find('input', {'name': '_token'})
                if csrf_input:
                    csrf_token = csrf_input.get('value')
                    logger.debug(f"✅ CSRF token: {csrf_token[:30]}...")
                else:
                    logger.error("❌ No se encontró CSRF token")
                    # Debug HTML content to understand why
                    logger.debug(f"HTML Preview: {html[:500]}")
                    return False
            
            # Paso 2: POST a /validar
            login_data = {
                '_token': csrf_token,
                'username': self.username,
                'password': self.password
            }
            
            async with self.session.post(
                f"{self.base_url}/validar",
                data=login_data,
                allow_redirects=False  # No seguir redirects
            ) as response:
                logger.debug(f"Status: {response.status}")
                
                # Login exitoso = 302 redirect a /menu
                if response.status == 302 and '/menu' in str(response.headers.get('Location', '')):
                    self.is_logged_in = True
                    logger.success("✅ Login exitoso!")
                    return True
                else:
                    logger.error(f"❌ Login fallido - Status: {response.status}")
                    return False
                    
        except Exception as e:
            logger.error(f"❌ Error en login: {e}")
            logger.error(traceback.format_exc())
            return False
    
    async def get_empleados(self) -> List[Dict[str, Any]]:
        """Obtener empleados descargando Excel completo"""
        try:
            if not await self.ensure_logged_in():
                return []
            
            logger.info("📥 Descargando Excel de usuarios...")
            
            # Descargar Excel
            async with self.session.get(f"{self.base_url}/usuario/ver/excel?") as response:
                if response.status != 200:
                    logger.error(f"❌ Error descargando Excel: {response.status}")
                    return []
                
                excel_data = await response.read()
                logger.info(f"✅ Excel descargado: {len(excel_data)} bytes")
            
            # Parsear Excel
            empleados = self._parse_excel(excel_data)
            
            logger.success(f"✅ {len(empleados)} empleados obtenidos del Excel")
            return empleados
            
        except Exception as e:
            logger.error(f"❌ Error obteniendo empleados: {e}")
            logger.error(traceback.format_exc())
            return []

    async def get_marcaciones(self, mes: int = None, anio: int = None, rut_filter: str = None, ruts_set: set = None) -> List[Dict[str, Any]]:
        """
        Obtener marcaciones (logs) de asistencia desde BioAlba via Excel Mensual.
        Si no se especifica mes/anio, usa el actual.
        Si rut_filter se provee, BioAlba filtra en servidor (reduce descarga drásticamente).
        Si ruts_set se provee, filtra durante el parsing (reduce objetos en RAM).
        """
        try:
            if not await self.ensure_logged_in():
                return []

            # Default a mes actual si no se provee
            now = datetime.now()
            if not mes: mes = now.month
            if not anio: anio = now.year

            month_str = f"{anio}-{mes:02d}"
            
            # Filtro por RUT: si se provee, BioAlba filtra en servidor
            rut_param = rut_filter or ""
            cache_buster = int(_t.time())
            
            # 1. VISITAR PAGINA DE FILTRO PRIMERO (Prime Session)
            # Esto es necesario para que el servidor genere el estado de sesión correcto
            filter_url = f"{self.base_url}/marcaciones?namerutcmp={rut_param}&bday-month={month_str}&_cb={cache_buster}"
            logger.info(f"🌍 Visitando filtro: {filter_url}")
            
            async with self.session.get(filter_url) as resp_filter:
                if resp_filter.status != 200:
                    logger.warning(f"⚠️ Error cargando filtro: {resp_filter.status}")

            # 2. DESCARGAR EXCEL
            filter_label = f" [RUT: {rut_filter}]" if rut_filter else " [TODOS]"
            logger.info(f"📥 Descargando logs de asistencia (Mes: {month_str}{filter_label})...")
            
            # URL verificada para descarga mensual
            download_url = f"{self.base_url}/marcacion/ver/excel?namerutcmp={rut_param}&bday-month={month_str}&_cb={cache_buster}"
            
            # Headers indispensables para evitar error 500
            headers = {
                "Referer": filter_url,
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            }
            
            
            async with self.session.get(download_url, headers=headers) as response:
                logger.debug(f"[BioAlba] Descarga status: {response.status}")
                if response.status != 200:
                    logger.error(f"❌ Error descargando logs: {response.status}")
                    return []

                content_type = response.headers.get("Content-Type", "")
                logger.debug(f"[BioAlba] Content-Type: {content_type}")
                if "text/html" in content_type:
                    logger.error("❌ El servidor devolvió HTML en lugar de Excel (Posible error de sesión)")
                    return []

                excel_data = await response.read()
                logger.info(f"✅ Excel de logs descargado: {len(excel_data)} bytes")

            # Parsear
            marcaciones = self._parse_excel_marcaciones(excel_data, ruts_filter=ruts_set)
            logger.success(f"✅ {len(marcaciones)} marcaciones obtenidas del Excel")
            return marcaciones

        except Exception as e:
            logger.error(f"❌ Error obteniendo marcaciones: {e}")
            logger.error(traceback.format_exc())
            return []

    def _parse_excel_marcaciones(self, excel_data: bytes, ruts_filter: set = None) -> List[Dict[str, Any]]:
        """Parsear archivo Excel de marcaciones (Rust/calamine o fallback openpyxl).
        Si ruts_filter se provee, solo crea dicts para RUTs en el set (filtro temprano de RAM).
        """
        try:
            t0 = _t.perf_counter()
            if _CALAMINE_AVAILABLE:
                wb = CalamineWorkbook.from_filelike(BytesIO(excel_data))
                sheet = wb.get_sheet_by_index(0)
                rows = list(sheet.to_python(skip_empty_area=False))
                rows = rows[1:]   # skip header
                logger.debug(f"⚡ calamine cargó {len(rows)} filas en {(_t.perf_counter()-t0)*1000:.0f}ms")
            else:
                wb = _openpyxl_load(BytesIO(excel_data), read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                wb.close()

            marcaciones = []
            skipped_by_filter = 0
            for row in rows:
                if not row or len(row) < 7:
                    continue

                rut_raw       = str(row[1]).strip() if row[1] is not None else ""
                fecha_hora_raw = str(row[3]).strip() if row[3] is not None else ""

                if not rut_raw or not fecha_hora_raw or fecha_hora_raw in ("", "None"):
                    continue

                rut_clean = self._format_rut(rut_raw)

                # ⚡ Filtro temprano: descartar ANTES de construir dict (ahorra RAM)
                if ruts_filter and rut_clean not in ruts_filter:
                    skipped_by_filter += 1
                    continue

                try:
                    dt_obj = datetime.strptime(fecha_hora_raw, "%d/%m/%y %H:%M:%S")
                    fecha_final = dt_obj.strftime("%Y-%m-%d %H:%M:%S")
                except ValueError:
                    try:
                        dt_obj = datetime.strptime(fecha_hora_raw, "%d/%m/%Y %H:%M:%S")
                        fecha_final = dt_obj.strftime("%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        logger.warning(f"⚠️ Formato fecha desconocido: {fecha_hora_raw}")
                        fecha_final = fecha_hora_raw

                marcaciones.append({
                    'rut':       rut_clean,
                    'fecha_hora': fecha_final,
                    'tipo':      str(row[4]).strip() if row[4] is not None else "Desconocido",
                    'equipo':    str(row[6]).strip() if len(row) > 6 and row[6] is not None else "Manual"
                })

            filter_msg = f" (filtro temprano descartó {skipped_by_filter} filas)" if skipped_by_filter else ""
            logger.info(f"✅ {len(marcaciones)} marcaciones parseadas en {(_t.perf_counter()-t0)*1000:.0f}ms{filter_msg}")
            return marcaciones
        except Exception as e:
            logger.error(f"Error parseando Excel marcaciones: {e}")
            return []
    
    def _parse_excel(self, excel_data: bytes) -> List[Dict[str, Any]]:
        """Parsear archivo Excel de empleados (Rust/calamine o fallback openpyxl)"""
        try:
            t0 = _t.perf_counter()

            if _CALAMINE_AVAILABLE:
                wb = CalamineWorkbook.from_filelike(BytesIO(excel_data))
                sheet = wb.get_sheet_by_index(0)
                all_rows = list(sheet.to_python(skip_empty_area=False))
                logger.info(f"⚡ calamine cargó {len(all_rows)} filas en {(_t.perf_counter()-t0)*1000:.0f}ms")
                rows_data = all_rows  # procesar desde idx 0, skip manual en loop
            else:
                wb = _openpyxl_load(BytesIO(excel_data), read_only=True)
                ws = wb.active
                rows_data = list(ws.iter_rows(min_row=1, values_only=True))
                wb.close()
                logger.info(f"📊 openpyxl cargó {len(rows_data)} filas")

            empleados = []

            for idx, row in enumerate(rows_data):
                if idx == 0: continue  # Skip header

                try:
                    if not row or len(row) < 1:
                        continue

                    rut_raw        = str(row[0]).strip() if row[0] is not None else ""
                    nombre_completo = str(row[1]).strip() if row[1] is not None else ""

                    if not rut_raw or rut_raw in ["None", "---", ""]:
                        continue

                    rut = self._format_rut(rut_raw)

                    if not nombre_completo or nombre_completo in ["---", "None"]:
                        nombre_completo = f"Usuario {rut}"

                    partes_nombre = self._parse_nombre(nombre_completo)

                    # Parsear fecha creacion
                    fecha_creacion_raw = str(row[6]).strip() if len(row) > 6 and row[6] is not None and str(row[6]) not in ['---', 'None'] else None
                    fecha_ingreso_formatted = None
                    if fecha_creacion_raw:
                        try:
                            date_part = fecha_creacion_raw.split(' ')[0]
                            parsed_dt = None
                            for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y', '%Y-%m-%d', '%H:%M']:
                                try:
                                    parsed_dt = datetime.strptime(date_part, fmt)
                                    break
                                except ValueError:
                                    continue
                            if parsed_dt and not (parsed_dt.year == 1900 and parsed_dt.month == 1 and parsed_dt.day == 1):
                                fecha_ingreso_formatted = parsed_dt.strftime("%Y-%m-%d")
                        except Exception:
                            pass

                    # Parsear género desde row[2] y normalizar
                    _GENERO_VACIO = {'---', 'NONE', 'N/A', 'NA', 'SIN DATO', 'NO ESPECIFICADO',
                                     'NO ESPECIFICA', 'NO INFORMADO', 'NOT SPECIFIED', 'OTROS', '-'}
                    genero_raw_str = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                    genero = None
                    if genero_raw_str and genero_raw_str.upper() not in _GENERO_VACIO:
                        g = genero_raw_str.upper()
                        if g in ('M', 'MASC', 'MASCULINO', 'HOMBRE', 'MALE'):
                            genero = 'Masculino'
                        elif g in ('F', 'FEM', 'FEMENINO', 'MUJER', 'FEMALE'):
                            genero = 'Femenino'
                        elif g in ('O', 'OTRO', 'OTHER', 'NB', 'NO BINARIO'):
                            genero = 'Otro'
                        else:
                            genero = genero_raw_str  # Conservar si es un valor corto reconocible

                    genero_raw = genero_raw_str or None  # Para el log

                    area_raw = str(row[8]).strip() if len(row) > 8 and row[8] is not None and str(row[8]) not in ['---', 'None', ''] else "SIN ASIGNAR"
                    if area_raw.upper() in ["SIN ASIGNAR", "POR ASIGNAR"]:
                        area_raw = "SIN ASIGNAR"

                    empleados.append({
                        'rut':              rut,
                        'nombre':           partes_nombre['nombre'],
                        'apellido_paterno': partes_nombre['apellido_paterno'],
                        'apellido_materno': partes_nombre['apellido_materno'],
                        'cargo':    str(row[7]).strip() if len(row) > 7 and row[7] is not None and str(row[7]) not in ['---', 'None', ''] else None,
                        'area':     area_raw,
                        'compania': str(row[5]).strip() if len(row) > 5 and row[5] is not None and str(row[5]) != '---' else 'Aguacol',
                        'email':    str(row[3]).strip() if len(row) > 3 and row[3] is not None and str(row[3]) not in ['---', 'None'] else None,
                        'telefono': str(row[4]).strip() if len(row) > 4 and row[4] is not None and str(row[4]) not in ['---', 'None'] else None,
                        'genero':   genero,
                        'fecha_ingreso': fecha_ingreso_formatted,
                        'activo': True
                    })


                except Exception as e:
                    logger.debug(f"Error en fila {idx}: {e}")
                    continue

            logger.success(f"✅ {len(empleados)} empleados parseados en {(_t.perf_counter()-t0)*1000:.0f}ms")
            return empleados

        except Exception as e:
            logger.error(f"❌ Error parseando Excel empleados: {e}")
            logger.error(traceback.format_exc())
            return []
    def _clean(self, text: str) -> str:

        """Limpiar texto"""
        if not text or text == "None":
            return ""
        return str(text).strip().replace('\n', ' ').replace('\r', '').replace('\t', ' ').strip()
    
    def _format_rut(self, rut: str) -> str:
        """Formatear RUT chileno para coincidir con la base de datos (solo números/K, sin guión)"""
        if not rut or rut == "None":
            return ""
        
        # Quitar todo excepto números y K/k
        rut_limpio = re.sub(r'[^0-9Kk]', '', str(rut))
        return rut_limpio.upper()
    
    def _parse_nombre(self, nombre_completo: str) -> Dict[str, str]:
        """
        Parsear nombre completo. 
        Formato BioAlba EXCEL: ApellidoPaterno ApellidoMaterno Nombres
        """
        if not nombre_completo or "No informado" in nombre_completo or nombre_completo == "None":
            return {'nombre': '', 'apellido_paterno': '', 'apellido_materno': ''}
        
        partes = str(nombre_completo).strip().split()
        
        if len(partes) >= 3:
            # PATERNO MATERNO NOMBRES...
            return {
                'apellido_paterno': partes[0],
                'apellido_materno': partes[1],
                'nombre': ' '.join(partes[2:])
            }
        elif len(partes) == 2:
            # PATERNO NOMBRES (Caso común cuando falta el materno)
            return {
                'apellido_paterno': partes[0],
                'apellido_materno': '',
                'nombre': partes[1]
            }
        else:
            # Solo una palabra -> Asumir que es el nombre o fallback
            return {
                'nombre': nombre_completo,
                'apellido_paterno': '',
                'apellido_materno': ''
            }
    
    async def test_connection(self) -> bool:
        """Probar conexión"""
        return await self.login()
