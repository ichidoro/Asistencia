import os
import pandas as pd
from datetime import datetime, date, timedelta
from io import BytesIO
from typing import List, Dict, Any, Optional
from loguru import logger

from backend.core.config import settings
from backend.services.asistencia_service import AsistenciaService

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image

class ReportService:
    def __init__(self, asistencia_service: AsistenciaService):
        self.asistencia_service = asistencia_service

    def _format_hhmmss(self, minutos) -> str:
        if minutos is None or minutos == "":
            return ""
        if minutos == 0:
            return "00:00:00"
        signo = "-" if minutos < 0 else ""
        total_secs = round(abs(minutos) * 60)
        h = total_secs // 3600
        m = (total_secs % 3600) // 60
        s = total_secs % 60
        return f"{signo}{h:02d}:{m:02d}:{s:02d}"

    def _get_saldo(self, extras_aprobados: int, deuda: int) -> str:
        saldo = (extras_aprobados or 0) - (deuda or 0)
        signo = "+" if saldo > 0 else ("-" if saldo < 0 else "")
        return f"{signo}{self._format_hhmmss(abs(saldo))}"

    async def generate_excel_report(self, fecha_inicio: str, fecha_fin: str, area: str = None, turno_id: int = None) -> BytesIO:
        """
        Genera el reporte Excel oficial de cierre o mensual para el periodo y area dados.
        """
        try:
            f_ini = date.fromisoformat(fecha_inicio)
            f_fin = date.fromisoformat(fecha_fin)
            rango_dias = [(f_ini + timedelta(days=i)) for i in range((f_fin - f_ini).days + 1)]
            
            # Obtener datos de la matriz del periodo
            matrix_data = await self.asistencia_service.get_matrix_data_with_projections(
                f_ini.month, f_ini.year, area=area, turno_id=turno_id, fecha_inicio_override=fecha_inicio, fecha_fin_override=fecha_fin
            )
            return await self._generate_excel_workbook(matrix_data, rango_dias, area)
        except Exception as e:
            logger.error(f"Error en generate_excel_report: {e}")
            raise

    async def generate_excel_custom_range(self, fecha_inicio: str, fecha_fin: str, area: str = None, turno_id: int = None) -> BytesIO:
        """
        Genera el reporte Excel para un rango personalizado.
        """
        try:
            f_ini = date.fromisoformat(fecha_inicio)
            f_fin = date.fromisoformat(fecha_fin)
            rango_dias = [(f_ini + timedelta(days=i)) for i in range((f_fin - f_ini).days + 1)]
            
            # Obtener datos usando el motor por rango
            matrix_data = await self.asistencia_service.get_matrix_data_by_range(
                fecha_inicio, fecha_fin, area=area, turno_id=turno_id
            )
            return await self._generate_excel_workbook(matrix_data, rango_dias, area)
        except Exception as e:
            logger.error(f"Error en generate_excel_custom_range: {e}")
            raise

    async def _generate_excel_workbook(self, matrix_data: dict, rango_dias: List[date], area: Optional[str]) -> BytesIO:
        try:
            empleados = matrix_data.get("empleados", [])
            emp_matrix = matrix_data.get("matrix", {})
            feriados = matrix_data.get("feriados", [])
            feriados_set = {f['fecha'] for f in feriados}
            
            # Evaluar bonos desde el backend
            from backend.services.bono_service import BonoService
            bono_srv = BonoService(self.asistencia_service.repository.db)
            bonos_eval = await bono_srv.evaluar_bonos_directo(
                empleados, 
                matrix_data.get("data", []), 
                matrix_data.get("justificaciones", []), 
                matrix_data.get("matrix", {})
            )
            
            todos_bonos = set()
            for emp_bonos in bonos_eval.values():
                todos_bonos.update(emp_bonos.keys())
            lista_bonos = sorted(list(todos_bonos))

            # 1. Calcular resumen completo de cada empleado (réplica exacta de la grilla analítica)
            rows_calc = []
            hay_bolsa = False
            
            for emp in empleados:
                emp_id = emp['id']
                es_bolsa = emp.get('tipo_programacion') == 'FLEXIBLE_BOLSA'
                if es_bolsa:
                    hay_bolsa = True
                    
                he_bruto = 0
                he_apr = 0
                he_rec = 0
                he_pend = 0
                he_compensado = 0
                d_tot = 0
                min_col = 0
                min_per = 0
                min_atr = 0
                min_sad = 0
                
                cnt_atr = 0
                cnt_sad = 0
                cnt_inas = 0
                cnt_esp = 0
                cnt_per = 0
                cnt_efectivos = 0
                
                acum_bolsa = 0
                excedido = False
                meta_min = 0
                
                # Cargar dias guardados de la matriz de asistencia
                dias_dict = emp_matrix.get(str(emp_id), {})
                if not dias_dict and emp_id in emp_matrix:
                    dias_dict = emp_matrix[emp_id]

                if es_bolsa:
                    meta_original = emp.get("meta_mensual_minutos")
                    if not meta_original:
                        meta_original = round((emp.get("meta_horas_semanales") or 0.0) * 60)
                    
                    dias_programados = 0
                    dias_justificados = 0
                    
                    for d in rango_dias:
                        f_str = f"{d.year}-{d.month:02d}-{d.day:02d}"
                        is_fer = f_str in feriados_set
                        di_check = dias_dict.get(f_str, {})
                        
                        day_db = d.weekday()
                        turno_dias = emp.get("turno_dias", {})
                        day_info = turno_dias.get(str(day_db), {}) or turno_dias.get(day_db, {})
                        is_structurally_libre = day_info.get("es_libre") == 1
                        
                        is_descanso = is_fer or is_structurally_libre or (di_check.get("estado") == 'LIBRE')
                        
                        if not is_descanso:
                            dias_programados += 1
                            estados_justificados = ['VACACIONES', 'LICENCIA', 'LIC_COMUN', 'LIC_MUTUAL', 'CUMPLEAÑOS', 'DUELO', 'PERMISO', 'NO NACIDO', 'DEFUNCION']
                            di_estado = di_check.get("estado", "") or ""
                            di_nomen = di_check.get("nomenclatura", "") or ""
                            is_justificado = (
                                any(ej in di_estado.upper() for ej in estados_justificados) or
                                (bool(di_nomen) and str(di_nomen).strip().upper() != 'DEOP')
                            )
                            
                            if is_justificado:
                                dias_justificados += 1
                                
                    # Evitar división por cero
                    if dias_programados > 0:
                        if dias_justificados > 0:
                            valor_turno_min = meta_original / dias_programados
                            meta_min = round(meta_original - (valor_turno_min * dias_justificados))
                        else:
                            meta_min = meta_original
                    else:
                        meta_min = meta_original
                        
                    if emp.get("meta_ajustada_minutos_descuento") and dias_justificados == 0:
                        meta_min = max(0, meta_min - emp.get("meta_ajustada_minutos_descuento"))
                
                acum_semanal = 0
                start_day_db = emp.get("primer_dia_semana_turno", 0)
                
                for d in rango_dias:
                    if d.weekday() == start_day_db:
                        acum_semanal = 0
                        
                    f_str = f"{d.year}-{d.month:02d}-{d.day:02d}"
                    di = dias_dict.get(f_str)
                    if not di:
                        continue
                        
                    trab = round((di.get("horas_trabajadas") or 0.0) * 60)
                    di_estado = di.get("estado") or ""
                    is_esp = di_estado in ['JORNADA_ESPECIAL', 'EXTRA', 'FERIADO Y JORNADA EXTRA', 'DÍA LIBRE Y JORNADA EXTRA']
                    
                    if not es_bolsa and not is_esp:
                        acum_semanal += trab
                        
                    di["_acumuladoSemanalSnap"] = acum_semanal
                    
                    if not is_esp and not es_bolsa:
                        tiene_condonacion = (di.get("deuda_condonada") or 0) > 0
                        net_deuda = 0 if tiene_condonacion else (di.get("minutos_deuda") or 0)
                        
                        raw_col = di.get("minutos_exceso_colacion") or 0
                        raw_per = di.get("minutos_permiso_personal_deuda") or 0
                        raw_atr = 0 if tiene_condonacion else (di.get("minutos_atraso") or 0)
                        raw_sad = 0 if tiene_condonacion else (di.get("minutos_salida_adelantada") or 0)
                        
                        raw_total = raw_col + raw_per + raw_atr + raw_sad
                        
                        day_col = 0
                        day_per = 0
                        day_atr = 0
                        day_sad = 0
                        
                        if net_deuda > 0 and raw_total > 0:
                            if net_deuda >= raw_total:
                                day_col = raw_col
                                day_per = raw_per
                                day_atr = raw_atr
                                day_sad = raw_sad
                            else:
                                factor = net_deuda / raw_total
                                day_col = raw_col * factor
                                day_per = raw_per * factor
                                day_atr = raw_atr * factor
                                day_sad = raw_sad * factor
                                
                        d_tot += net_deuda
                        min_col += day_col
                        min_per += day_per
                        min_atr += day_atr
                        min_sad += day_sad
                        
                        if (di.get("minutos_atraso") or 0) > 0 and not tiene_condonacion:
                            cnt_atr += 1
                        if (di.get("minutos_salida_adelantada") or 0) > 0 and not tiene_condonacion:
                            cnt_sad += 1
                        if di.get("tiene_permiso_hora") or di.get("permiso_activo"):
                            cnt_per += 1
                            
                    if di_estado == 'INASISTENCIA':
                        cnt_inas += 1
                    if is_esp:
                        cnt_esp += 1
                    if di.get("hora_entrada_real") and not is_esp and di_estado not in ['LIBRE', 'FERIADO', 'INASISTENCIA']:
                        cnt_efectivos += 1
                        
                    # HE calculations
                    if not is_esp:
                        if es_bolsa:
                            snap_antes = acum_bolsa
                            acum_bolsa += trab
                            di["_acumuladoBolsaSnapPrev"] = snap_antes
                            di["_acumuladoBolsaSnap"] = acum_bolsa
                            di["_metaMinBolsa"] = meta_min
                            
                            if excedido:
                                he_bruto += trab
                            elif acum_bolsa > meta_min and trab > 0:
                                he_bruto += (acum_bolsa - meta_min)
                                excedido = True
                        else:
                            he_bruto += max(di.get("minutos_extra_bruto") or 0, di.get("minutos_extra_autorizados") or 0)
                    elif es_bolsa:
                        di["_acumuladoBolsaSnapPrev"] = acum_bolsa
                        di["_acumuladoBolsaSnap"] = acum_bolsa
                        di["_metaMinBolsa"] = meta_min
                        
                    if not is_esp:
                        if di.get("estado_he") == 'APROBADO':
                            he_apr += (di.get("minutos_extra_autorizados") or 0)
                        elif di.get("estado_he") == 'RECHAZADO':
                            he_rec += (di.get("minutos_extra_bruto") or 0)
                        elif (di.get("minutos_extra_bruto") or 0) > 0:
                            he_pend += (di.get("minutos_extra_bruto") or 0)
                    # FIX: he_compensado se acumula para TODOS los días (incluidos JORNADA_ESPECIAL),
                    # igual que en el frontend (marcaciones_ui.js línea 4266)
                    he_compensado += (di.get("minutos_compensados_he") or 0)
                            
                # FIX: Redondeo con 4 decimales para coincidir con el frontend
                he_bruto = round(he_bruto, 4)
                he_apr = round(he_apr, 4)
                he_rec = round(he_rec, 4)
                he_pend = round(he_pend, 4)
                
                saldo = he_apr - d_tot - he_compensado
                saldo_meta = (acum_bolsa - meta_min) if es_bolsa else None
                
                # DEBUG: Log de auditoría para detectar discrepancia Excel vs Grilla
                emp_nombre = f"{emp['apellido_paterno']} {emp.get('apellido_materno', '')} {emp['nombre']}".strip()
                logger.info(
                    f"📊 [EXCEL_DEBUG] {emp_nombre}: "
                    f"he_apr={he_apr} d_tot={d_tot} he_comp={he_compensado} "
                    f"saldo={saldo} ({self._format_hhmmss(saldo)}) | "
                    f"he_bruto={he_bruto} he_rec={he_rec} he_pend={he_pend} "
                    f"dias_procesados={sum(1 for d in rango_dias if dias_dict.get(f'{d.year}-{d.month:02d}-{d.day:02d}'))}"
                )
                
                rows_calc.append({
                    "emp_id": emp_id,
                    "nombre": f"{emp['apellido_paterno']} {emp.get('apellido_materno', '')} {emp['nombre']}".strip().replace('  ', ' '),
                    "rut": emp.get("rut", ""),
                    "activo": "SÍ" if emp.get("activo", 1) else "NO",
                    "esBolsa": es_bolsa,
                    "he_bruto": he_bruto,
                    "he_apr": he_apr,
                    "he_rec": he_rec,
                    "he_pend": he_pend,
                    "d_tot": d_tot,
                    "min_col": round(min_col),
                    "min_per": round(min_per),
                    "min_atr": round(min_atr),
                    "min_sad": round(min_sad),
                    "cnt_atr": cnt_atr,
                    "cnt_sad": cnt_sad,
                    "cnt_inas": cnt_inas,
                    "cnt_esp": cnt_esp,
                    "cnt_per": cnt_per,
                    "cnt_efectivos": cnt_efectivos,
                    "saldo": saldo,
                    "metaMin": meta_min,
                    "acumBolsa": acum_bolsa,
                    "saldoMeta": saldo_meta,
                    "bonos": bonos_eval.get(emp_id) or bonos_eval.get(str(emp_id), {})
                })

            # 2. Inicializar openpyxl Workbook
            wb = Workbook()
            
            # Paleta de estilos corporativos
            fill_empleado = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            fill_bonos = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
            fill_incidencias = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            fill_he = PatternFill(start_color="CFE2FF", end_color="CFE2FF", fill_type="solid")
            fill_deudas = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
            fill_saldo = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
            fill_bolsa = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
            fill_dias = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")
            
            font_header_top = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
            font_header_sub = Font(name="Segoe UI", size=9, bold=True, color="475569")
            
            borde_fino = Border(
                left=Side(style='thin', color="DDDDDD"), 
                right=Side(style='thin', color="DDDDDD"), 
                top=Side(style='thin', color="DDDDDD"), 
                bottom=Side(style='thin', color="DDDDDD")
            )
            borde_totales = Border(
                top=Side(style='thin', color="000000"),
                bottom=Side(style='double', color="000000"),
                left=Side(style='thin', color="DDDDDD"),
                right=Side(style='thin', color="DDDDDD")
            )
            
            dias_espanol = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
            monthNamesShort = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
            
            # Mapear columnas dinámicas de la hoja
            columns_def = [
                ("Empleado", "Identificador", "Identificador"),
                ("Empleado", "RUT", "RUT"),
                ("Empleado", "Empleado", "Empleado"),
                ("Empleado", "Activo", "Activo")
            ]
            for b in lista_bonos:
                columns_def.append(("Bonos", b, f"bono_{b}"))
                
            columns_def.extend([
                ("Incidencias", "PERM", "cnt_per"),
                ("Incidencias", "ATR", "cnt_atr"),
                ("Incidencias", "S.ADL", "cnt_sad"),
                ("Incidencias", "INA", "cnt_inas"),
                ("Incidencias", "J.ESP", "cnt_esp"),
                ("Incidencias", "TOT", "tot_incidencias")
            ])
            columns_def.extend([
                ("Horas Extra", "PEND", "he_pend"),
                ("Horas Extra", "APR", "he_apr"),
                ("Horas Extra", "RECH", "he_rech"),
                ("Horas Extra", "TOT", "he_tot")
            ])
            columns_def.extend([
                ("Tiempo No Trabajado", "COL", "min_col"),
                ("Tiempo No Trabajado", "PER", "min_per"),
                ("Tiempo No Trabajado", "ATR", "min_atr"),
                ("Tiempo No Trabajado", "S.ADL", "min_sad"),
                ("Tiempo No Trabajado", "TOT", "d_tot")
            ])
            columns_def.append(("Saldo Neto", "Saldo Neto", "saldo_neto"))
            
            if hay_bolsa:
                columns_def.extend([
                    ("Bolsa Flexible", "META", "meta_min"),
                    ("Bolsa Flexible", "ACUM.", "acum_bolsa"),
                    ("Bolsa Flexible", "BALANCE", "saldo_meta")
                ])
                
            for d in rango_dias:
                fecha_str = f"{d.year}-{d.month:02d}-{d.day:02d}"
                dia_semana_str = dias_espanol[d.weekday()]
                header_day = f"{d.day:02d}-{monthNamesShort[d.month-1]}\n{dia_semana_str[:3].upper()}"
                columns_def.append(("Días del período", header_day, f"day_{fecha_str}"))

            group_fill_map = {
                "Empleado": fill_empleado,
                "Bonos": fill_bonos,
                "Incidencias": fill_incidencias,
                "Horas Extra": fill_he,
                "Tiempo No Trabajado": fill_deudas,
                "Saldo Neto": fill_saldo,
                "Bolsa Flexible": fill_bolsa,
                "Días del período": fill_dias
            }

            pestañas = [
                ("Conceptos", "conceptos"),
                ("Horas Reales", "horas"),
                ("Colacion", "colacion"),
                ("Permisos", "permisos"),
                ("Horas Extras", "he"),
                ("Acumulado", "acumulado")
            ]
            
            logo_path = os.path.join(settings.BASE_DIR, "frontend", "assets", "img", "logo.jpg")

            for sheet_title, view_mode in pestañas:
                ws = wb.create_sheet(title=sheet_title)
                ws.sheet_view.showGridLines = True
                
                # Membrete y Logo corporativo
                if os.path.exists(logo_path):
                    try:
                        img_header = Image(logo_path)
                        img_header.height = 75
                        img_header.width = 170
                        ws.add_image(img_header, 'A1')
                    except Exception as ex:
                        logger.warning(f"No se pudo cargar el logo en Excel: {ex}")
                        
                # Título principal
                ws.merge_cells('C2:H2')
                title_cell = ws['C2']
                title_cell.value = f"REPORTE OFICIAL DE ASISTENCIA - {sheet_title.upper()}"
                title_cell.font = Font(name="Segoe UI", size=14, bold=True, color="003366")
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Subtítulo
                ws.merge_cells('C3:H3')
                sub_cell = ws['C3']
                sub_cell.value = "DEPARTAMENTO DE RECURSOS HUMANOS"
                sub_cell.font = Font(name="Segoe UI", size=10, bold=True, color="666666")
                sub_cell.alignment = Alignment(horizontal='center', vertical='top')
                
                # Metadata del reporte
                ws['I1'] = "Fecha Emisión:"
                ws['I1'].font = Font(name="Segoe UI", size=9, bold=True, color="555555")
                ws['I1'].alignment = Alignment(horizontal='right')
                ws['J1'] = datetime.now().strftime('%d/%m/%Y %H:%M')
                ws['J1'].font = Font(name="Segoe UI", size=9, color="333333")
                ws['J1'].alignment = Alignment(horizontal='left')
                
                ws['I2'] = "Período Reportado:"
                ws['I2'].font = Font(name="Segoe UI", size=9, bold=True, color="555555")
                ws['I2'].alignment = Alignment(horizontal='right')
                ws['J2'] = f"{rango_dias[0].strftime('%d/%m/%Y')} al {rango_dias[-1].strftime('%d/%m/%Y')}"
                ws['J2'].font = Font(name="Segoe UI", size=9, color="333333")
                ws['J2'].alignment = Alignment(horizontal='left')
                
                ws['I3'] = "Sistema:"
                ws['I3'].font = Font(name="Segoe UI", size=9, bold=True, color="555555")
                ws['I3'].alignment = Alignment(horizontal='right')
                ws['J3'] = "Aguacol Asistencia"
                ws['J3'].font = Font(name="Segoe UI", size=9, color="333333")
                ws['J3'].alignment = Alignment(horizontal='left')
                
                ws['I4'] = "Área Reportada:"
                ws['I4'].font = Font(name="Segoe UI", size=9, bold=True, color="555555")
                ws['I4'].alignment = Alignment(horizontal='right')
                ws['J4'] = str(area).upper() if area else "TODAS LAS ÁREAS"
                ws['J4'].font = Font(name="Segoe UI", size=9, color="333333")
                ws['J4'].alignment = Alignment(horizontal='left')
                
                # Escribir cabeceras Bento (top y sub)
                for col_idx, (group_name, col_name, _) in enumerate(columns_def, start=1):
                    fill = group_fill_map.get(group_name, fill_empleado)
                    
                    cell_top = ws.cell(row=6, column=col_idx, value=group_name)
                    cell_top.fill = fill
                    cell_top.font = font_header_top
                    cell_top.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell_top.border = borde_fino
                    
                    cell_sub = ws.cell(row=7, column=col_idx, value=col_name)
                    cell_sub.fill = fill
                    cell_sub.font = font_header_sub
                    cell_sub.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell_sub.border = borde_fino
                    
                # Realizar combinaciones del top-header
                start_col = 1
                current_group = columns_def[0][0]
                for col_idx, (group_name, _, _) in enumerate(columns_def, start=1):
                    if group_name != current_group:
                        if col_idx - 1 > start_col:
                            ws.merge_cells(start_row=6, start_column=start_col, end_row=6, end_column=col_idx - 1)
                        start_col = col_idx
                        current_group = group_name
                if len(columns_def) > start_col:
                    ws.merge_cells(start_row=6, start_column=start_col, end_row=6, end_column=len(columns_def))
                    
                ws.row_dimensions[6].height = 25
                ws.row_dimensions[7].height = 25
                
                # Congelar paneles: Columnas A-D inmovilizadas y filas 1-7 inmovilizadas
                ws.freeze_panes = 'E8'
                
                # Escribir las filas de datos de empleados
                current_row = 8
                for r in rows_calc:
                    emp_id = r["emp_id"]
                    dias_dict = emp_matrix.get(str(emp_id), {})
                    if not dias_dict and emp_id in emp_matrix:
                        dias_dict = emp_matrix[emp_id]
                        
                    for col_idx, (_, _, key) in enumerate(columns_def, start=1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.border = borde_fino
                        cell.font = Font(name="Segoe UI", size=9)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        val = ""
                        if key == "Identificador":
                            val = emp_id
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif key == "RUT":
                            val = r["rut"]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif key == "Empleado":
                            val = r["nombre"]
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        elif key == "Activo":
                            val = r["activo"]
                        # Bonos
                        elif key.startswith("bono_"):
                            b_name = key.replace("bono_", "")
                            b_res = r["bonos"].get(b_name)
                            if not b_res or b_res.get("aplica") is False:
                                val = "-"
                            else:
                                val = "SÍ" if b_res.get("califica") else "NO"
                                if val == "SÍ":
                                    cell.font = Font(name="Segoe UI", size=9, color="008000", bold=True)
                                else:
                                    cell.font = Font(name="Segoe UI", size=9, color="FF0000")
                        # Incidencias
                        elif key == "cnt_per": val = r["cnt_per"] if r["cnt_per"] > 0 else ""
                        elif key == "cnt_atr": val = r["cnt_atr"] if r["cnt_atr"] > 0 else ""
                        elif key == "cnt_sad": val = r["cnt_sad"] if r["cnt_sad"] > 0 else ""
                        elif key == "cnt_inas": 
                            val = r["cnt_inas"] if r["cnt_inas"] > 0 else ""
                            if r["cnt_inas"] > 0:
                                cell.font = Font(name="Segoe UI", size=9, color="FF0000", bold=True)
                        elif key == "cnt_esp": val = f"{r['cnt_esp']} *" if r["cnt_esp"] > 0 else ""
                        elif key == "tot_incidencias":
                            tot_i = r["cnt_per"] + r["cnt_atr"] + r["cnt_sad"] + r["cnt_inas"] + r["cnt_esp"]
                            val = tot_i if tot_i > 0 else ""
                        # HE
                        elif key == "he_pend": val = self._format_hhmmss(r["he_pend"])
                        elif key == "he_apr": val = self._format_hhmmss(r["he_apr"])
                        elif key == "he_rech": val = self._format_hhmmss(r["he_rec"])
                        elif key == "he_tot": val = self._format_hhmmss(r["he_bruto"])
                        # Deudas
                        elif key == "min_col": val = self._format_hhmmss(r["min_col"])
                        elif key == "min_per": val = self._format_hhmmss(r["min_per"])
                        elif key == "min_atr": val = self._format_hhmmss(r["min_atr"])
                        elif key == "min_sad": val = self._format_hhmmss(r["min_sad"])
                        elif key == "d_tot": val = self._format_hhmmss(r["d_tot"])
                        # Saldo Neto — FIX: usar el saldo pre-calculado que incluye he_compensado
                        # Antes: self._get_saldo(he_apr, d_tot) → ignoraba he_compensado
                        # Ahora: r["saldo"] = he_apr - d_tot - he_compensado (idéntico a la grilla)
                        elif key == "saldo_neto":
                            saldo_val = r["saldo"]
                            signo = "+" if saldo_val > 0 else ("-" if saldo_val < 0 else "")
                            val = f"{signo}{self._format_hhmmss(abs(saldo_val))}"
                            if saldo_val > 0:
                                cell.font = Font(name="Segoe UI", size=9, color="008000", bold=True)
                            elif saldo_val < 0:
                                cell.font = Font(name="Segoe UI", size=9, color="FF0000", bold=True)
                        # Bolsa Flexible
                        elif key == "meta_min": val = self._format_hhmmss(r["metaMin"]) if r["esBolsa"] else ""
                        elif key == "acum_bolsa": val = self._format_hhmmss(r["acumBolsa"]) if r["esBolsa"] else ""
                        elif key == "saldo_meta":
                            if r["esBolsa"]:
                                sm = r["saldoMeta"]
                                sm_sign = "+" if sm > 0 else ("-" if sm < 0 else "")
                                val = f"{sm_sign}{self._format_hhmmss(abs(sm))}"
                                if sm >= 0:
                                    cell.font = Font(name="Segoe UI", size=9, color="008000", bold=True)
                                else:
                                    cell.font = Font(name="Segoe UI", size=9, color="FF0000", bold=True)
                        # Días individuales
                        elif key.startswith("day_"):
                            f_str = key.replace("day_", "")
                            di = dias_dict.get(f_str)
                            
                            if di:
                                di_estado = di.get("estado") or ""
                                trab = round((di.get("horas_trabajadas") or 0.0) * 60)
                                deuda = di.get("minutos_deuda") or 0
                                
                                # Pestaña Conceptos
                                if view_mode == "conceptos":
                                    sigla = di_estado
                                    if di_estado == 'OK': 
                                        sigla = 'OK'
                                        cell.font = Font(name="Segoe UI", size=9, color="008000", bold=True)
                                    elif di_estado == 'ATRASO': 
                                        sigla = 'ATR'
                                        cell.font = Font(name="Segoe UI", size=9, color="FF8C00", bold=True)
                                    elif di_estado == 'SALIDA_ADELANTADA': 
                                        sigla = 'SAD'
                                        cell.font = Font(name="Segoe UI", size=9, color="800080", bold=True)
                                    elif di_estado == 'ATR_SAD': 
                                        sigla = 'ATR-SAD'
                                        cell.font = Font(name="Segoe UI", size=9, color="FF00FF", bold=True)
                                    elif di_estado == 'INASISTENCIA' or 'FALTA' in di_estado: 
                                        sigla = 'INA'
                                        cell.font = Font(name="Segoe UI", size=9, color="FF0000", bold=True)
                                    elif di_estado == 'LIBRE': 
                                        sigla = 'LIB'
                                        cell.font = Font(name="Segoe UI", size=9, color="808080")
                                    elif di_estado == 'FERIADO': 
                                        sigla = 'FES'
                                        cell.font = Font(name="Segoe UI", size=9, color="FFD700", bold=True)
                                    elif di.get("nomenclatura") or di_estado in ['VACACIONES', 'LICENCIA', 'LCM', 'LMU', 'CUM', 'DUE']: 
                                        sigla = di.get("nomenclatura") or di_estado[:3].upper()
                                        cell.font = Font(name="Segoe UI", size=9, color="0000FF", bold=True)
                                    val = sigla
                                    
                                # Pestaña Horas
                                elif view_mode == "horas":
                                    if di_estado in ['OK', 'ATRASO', 'SALIDA_ADELANTADA', 'ATR_SAD', 'JORNADA_ESPECIAL', 'EXTRA', 'EN_CURSO']:
                                        if di_estado == 'EN_CURSO':
                                            val = ">>>"
                                        else:
                                            hrs = self._format_hhmmss(trab)
                                            if not r["esBolsa"] and deuda > 0:
                                                val = f"{hrs}\nDEUDA {self._format_hhmmss(deuda)}"
                                                cell.font = Font(name="Segoe UI", size=8, color="FF0000")
                                            else:
                                                val = hrs
                                                
                                # Pestaña Colación
                                elif view_mode == "colacion":
                                    if di_estado in ['OK', 'ATRASO', 'SALIDA_ADELANTADA', 'ATR_SAD', 'JORNADA_ESPECIAL', 'EXTRA', 'EN_CURSO']:
                                        if di.get("hora_entrada_real") and di.get("hora_salida_real"):
                                            col_aplicado = di.get("minutos_colacion") or 0
                                            exceso = di.get("minutos_exceso_colacion") or 0
                                            col_str = self._format_hhmmss(col_aplicado)
                                            if exceso > 0:
                                                val = f"{col_str}\nDEUDA {self._format_hhmmss(exceso)}"
                                                cell.font = Font(name="Segoe UI", size=8, color="FF0000")
                                            elif di.get("minutos_colacion_real") == 0 and col_aplicado > 0:
                                                val = f"{col_str}\nAUTO"
                                            else:
                                                val = col_str
                                                
                                # Pestaña Permisos
                                elif view_mode == "permisos":
                                    m_perm = di.get("minutos_permisos_detectados") or 0
                                    m_deuda = di.get("minutos_permiso_personal_deuda") or 0
                                    if m_perm > 0:
                                        val = self._format_hhmmss(m_perm)
                                        cell.font = Font(name="Segoe UI", size=9, color="0000FF")
                                    elif di.get("tiene_permiso_hora") or di.get("permiso_activo"):
                                        mins = m_deuda if m_deuda > 0 else round((di.get("horas_teoricas") or 0.0) * 60)
                                        if mins > 0:
                                            val = self._format_hhmmss(mins)
                                        else:
                                            val = "PER"
                                        cell.font = Font(name="Segoe UI", size=9, color="FF8C00")
                                    elif di_estado in ['OK', 'ATRASO', 'SALIDA_ADELANTADA', 'ATR_SAD', 'JORNADA_ESPECIAL', 'EXTRA', 'EN_CURSO']:
                                        val = "—"
                                        
                                # Pestaña Horas Extras
                                elif view_mode == "he":
                                    if r["esBolsa"]:
                                        snap_hoy = di.get("_acumuladoBolsaSnap") or 0
                                        snap_ayer = di.get("_acumuladoBolsaSnapPrev") or 0
                                        meta_min_bolsa = di.get("_metaMinBolsa") or 0
                                        trab_hoy = round((di.get("horas_trabajadas") or 0.0) * 60)
                                        if snap_hoy > meta_min_bolsa and trab_hoy > 0:
                                            he_este = trab_hoy if snap_ayer >= meta_min_bolsa else (snap_hoy - meta_min_bolsa)
                                            if he_este > 0:
                                                val = f"+{self._format_hhmmss(he_este)}"
                                                cell.font = Font(name="Segoe UI", size=9, color="008000", bold=True)
                                    else:
                                        he_bruto_d = di.get("minutos_extra_bruto") or 0
                                        deuda_d = di.get("minutos_deuda") or 0
                                        txt = ""
                                        if he_bruto_d > 0:
                                            txt += f"+{self._format_hhmmss(he_bruto_d)}"
                                        if deuda_d > 0:
                                            txt += f"\n-{self._format_hhmmss(deuda_d)}"
                                        val = txt.strip()
                                        if "+" in val and "-" in val:
                                            cell.font = Font(name="Segoe UI", size=8)
                                        elif "+" in val:
                                            cell.font = Font(name="Segoe UI", size=9, color="008000", bold=True)
                                        elif "-" in val:
                                            cell.font = Font(name="Segoe UI", size=9, color="FF0000")
                                            
                                # Pestaña Acumulado
                                elif view_mode == "acumulado":
                                    if r["esBolsa"]:
                                        if di_estado in ['OK', 'ATRASO', 'SALIDA_ADELANTADA', 'ATR_SAD', 'JORNADA_ESPECIAL', 'EXTRA', 'EN_CURSO']:
                                            snap = di.get("_acumuladoBolsaSnap") or 0
                                            snap_ayer = di.get("_acumuladoBolsaSnapPrev") or 0
                                            meta_min_bolsa = di.get("_metaMinBolsa") or 0
                                            if snap > 0:
                                                cruza = snap > meta_min_bolsa and snap_ayer < meta_min_bolsa
                                                txt = self._format_hhmmss(snap)
                                                val = f"{txt}\n*META" if cruza else txt
                                                if snap > meta_min_bolsa:
                                                    cell.font = Font(name="Segoe UI", size=9, color="008000", bold=True)
                                                else:
                                                    cell.font = Font(name="Segoe UI", size=9, color="0000FF")
                                    else:
                                        if di_estado in ['OK', 'ATRASO', 'SALIDA_ADELANTADA', 'ATR_SAD', 'JORNADA_ESPECIAL', 'EXTRA', 'EN_CURSO'] and di.get("horas_trabajadas", 0) > 0:
                                            snap = di.get("_acumuladoSemanalSnap") or 0
                                            val = self._format_hhmmss(snap) if snap > 0 else "—"
                                
                                if not val and view_mode == "conceptos":
                                    if di.get('es_fin_de_semana'):
                                        val = "---"
                            else:
                                es_libre_puro = False
                                if emp.get("turno_dias") and f_str:
                                    py_day = d.weekday()
                                    day_info = emp.get("turno_dias", {}).get(py_day, {})
                                    if day_info.get("es_libre") == 1:
                                        es_libre_puro = True
                                        
                                if f_str in feriados_set:
                                    if view_mode == "conceptos":
                                        val = "FES"
                                        cell.font = Font(name="Segoe UI", size=9, color="FFD700", bold=True)
                                elif es_libre_puro:
                                    if view_mode == "conceptos":
                                        val = "LIB"
                                        cell.font = Font(name="Segoe UI", size=9, color="808080")
                                        
                        cell.value = val
                        
                    current_row += 1
                    
                # 5. Fila de TOTALES
                total_row_idx = current_row
                for col_idx, (_, col_name, key) in enumerate(columns_def, start=1):
                    cell = ws.cell(row=total_row_idx, column=col_idx)
                    cell.border = borde_totales
                    cell.font = Font(name="Segoe UI", size=9, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    val = ""
                    if key == "Identificador":
                        val = "TOTALES"
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    elif key in ["RUT", "Empleado", "Activo"] or key.startswith("bono_"):
                        val = ""
                    elif key == "cnt_per":
                        s_val = sum(r["cnt_per"] for r in rows_calc)
                        val = s_val if s_val > 0 else ""
                    elif key == "cnt_atr":
                        s_val = sum(r["cnt_atr"] for r in rows_calc)
                        val = s_val if s_val > 0 else ""
                    elif key == "cnt_sad":
                        s_val = sum(r["cnt_sad"] for r in rows_calc)
                        val = s_val if s_val > 0 else ""
                    elif key == "cnt_inas":
                        s_val = sum(r["cnt_inas"] for r in rows_calc)
                        val = s_val if s_val > 0 else ""
                    elif key == "cnt_esp":
                        s_val = sum(r["cnt_esp"] for r in rows_calc)
                        val = s_val if s_val > 0 else ""
                    elif key == "tot_incidencias":
                        s_val = sum(r["cnt_per"] + r["cnt_atr"] + r["cnt_sad"] + r["cnt_inas"] + r["cnt_esp"] for r in rows_calc)
                        val = s_val if s_val > 0 else ""
                    elif key == "he_pend": val = self._format_hhmmss(sum(r["he_pend"] for r in rows_calc))
                    elif key == "he_apr": val = self._format_hhmmss(sum(r["he_apr"] for r in rows_calc))
                    elif key == "he_rech": val = self._format_hhmmss(sum(r["he_rec"] for r in rows_calc))
                    elif key == "he_tot": val = self._format_hhmmss(sum(r["he_bruto"] for r in rows_calc))
                    elif key == "min_col": val = self._format_hhmmss(sum(r["min_col"] for r in rows_calc))
                    elif key == "min_per": val = self._format_hhmmss(sum(r["min_per"] for r in rows_calc))
                    elif key == "min_atr": val = self._format_hhmmss(sum(r["min_atr"] for r in rows_calc))
                    elif key == "min_sad": val = self._format_hhmmss(sum(r["min_sad"] for r in rows_calc))
                    elif key == "d_tot": val = self._format_hhmmss(sum(r["d_tot"] for r in rows_calc))
                    elif key == "saldo_neto":
                        # FIX: usar r["saldo"] pre-calculado que incluye he_compensado
                        s_tot = sum(r["saldo"] for r in rows_calc)
                        signo = "+" if s_tot > 0 else ("-" if s_tot < 0 else "")
                        val = f"{signo}{self._format_hhmmss(abs(s_tot))}"
                        if s_tot > 0:
                            cell.font = Font(name="Segoe UI", size=9, color="008000", bold=True)
                        elif s_tot < 0:
                            cell.font = Font(name="Segoe UI", size=9, color="FF0000", bold=True)
                    elif key in ["meta_min", "acum_bolsa", "saldo_meta"]:
                        val = ""
                        
                    cell.value = val
                    
                ws.row_dimensions[total_row_idx].height = 22
                
                # 6. Leyendas y pie de pagina
                foot_start = total_row_idx + 2
                ws.merge_cells(f"B{foot_start}:J{foot_start}")
                priv_cell = ws.cell(row=foot_start, column=2, value="AVISO DE CONFIDENCIALIDAD: Este documento y su contenido son de uso estrictamente interno y confidencial.")
                priv_cell.font = Font(name="Segoe UI", size=9, bold=True, color="AA0000")
                
                foot_start += 1
                ws.merge_cells(f"B{foot_start}:J{foot_start}")
                foot_cell = ws.cell(row=foot_start, column=2, value=f"Generado automáticamente por el Sistema de Asistencia de Aguacol S.A. - {datetime.now().strftime('%d/%m/%Y')}")
                foot_cell.font = Font(name="Segoe UI", size=9, italic=True, color="888888")
                
                # Autoajustar el ancho de las columnas
                for col in ws.columns:
                    max_length = 0
                    column_letter = col[0].column_letter
                    for cell in col:
                        if cell.row >= total_row_idx + 2:
                            continue
                        try:
                            if cell.value:
                                lines = str(cell.value).split('\n')
                                longest_line = max(len(l) for l in lines)
                                if longest_line > max_length:
                                    max_length = longest_line
                        except:
                            pass
                    adjusted_width = (max_length + 3)
                    if adjusted_width < 8: adjusted_width = 8
                    if adjusted_width > 45: adjusted_width = 45
                    ws.column_dimensions[column_letter].width = adjusted_width

            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
                
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error generando reporte Excel réplica: {e}")
            raise
